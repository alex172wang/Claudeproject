"""
L5偏差日志报告生成模块

提供PDF和Excel格式的偏差分析报告生成功能
"""

import io
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from decimal import Decimal

from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone

from .models import DeviationLog
from .l5_analytics import DeviationAnalytics, IntuitionTracker


class DeviationReportGenerator:
    """
    偏差报告生成器基类
    """

    def __init__(self, user=None, start_date=None, end_date=None):
        """
        初始化报告生成器

        Args:
            user: 指定用户（None表示所有用户）
            start_date: 报告开始日期
            end_date: 报告结束日期
        """
        self.user = user
        self.start_date = start_date or (timezone.now() - timedelta(days=30))
        self.end_date = end_date or timezone.now()

        # 初始化分析器
        self.analytics = DeviationAnalytics(
            user=user,
            start_date=self.start_date,
            end_date=self.end_date
        )

        if user:
            self.intuition_tracker = IntuitionTracker(user)
        else:
            self.intuition_tracker = None

    def _get_report_metadata(self) -> Dict[str, Any]:
        """获取报告元数据"""
        return {
            'report_title': 'L5偏差日志分析报告',
            'generated_at': timezone.now(),
            'date_range': {
                'start': self.start_date.strftime('%Y-%m-%d'),
                'end': self.end_date.strftime('%Y-%m-%d'),
            },
            'user': str(self.user) if self.user else '所有用户',
        }


class ExcelDeviationReport(DeviationReportGenerator):
    """
    Excel偏差报告生成器
    """

    def generate(self) -> io.BytesIO:
        """
        生成Excel格式的偏差报告

        Returns:
            包含Excel文件的BytesIO对象
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import PieChart, Reference, BarChart
        except ImportError:
            raise ImportError("请安装 pandas 和 openpyxl: pip install pandas openpyxl")

        # 创建工作簿
        wb = Workbook()

        # === 第1个工作表：概览 ===
        ws_overview = wb.active
        ws_overview.title = "概览"
        self._create_overview_sheet(ws_overview)

        # === 第2个工作表：偏差明细 ===
        ws_detail = wb.create_sheet("偏差明细")
        self._create_detail_sheet(ws_detail)

        # === 第3个工作表：类型统计 ===
        ws_type = wb.create_sheet("类型统计")
        self._create_type_stats_sheet(ws_type)

        # === 第4个工作表：直觉分析 ===
        if self.intuition_tracker:
            ws_intuition = wb.create_sheet("直觉分析")
            self._create_intuition_sheet(ws_intuition)

        # 保存到BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    def _create_overview_sheet(self, ws):
        """创建概览工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # 获取统计数据
        overview = self.analytics.get_overview_stats()

        # 标题
        ws['A1'] = 'L5偏差日志分析报告'
        ws['A1'].font = Font(size=20, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # 报告信息
        ws['A3'] = '报告期间:'
        ws['B3'] = f"{overview['date_range']['start']} 至 {overview['date_range']['end']}"
        ws['A4'] = '生成时间:'
        ws['B4'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = '用户:'
        ws['B5'] = str(self.user) if self.user else '所有用户'

        # 核心指标
        row = 8
        ws[f'A{row}'] = '核心指标'
        ws[f'A{row}'].font = Font(size=14, bold=True)

        row += 1
        metrics = [
            ('总偏差数', overview['total_deviations']),
            ('已验证数', overview['verification']['verified_count']),
            ('判断正确数', overview['verification']['correct_count']),
            ('直觉准确率', f"{overview['intuition_accuracy']['accuracy_rate']}%"),
            ('盈亏差异总计', f"¥{overview['pnl_stats']['total_pnl_diff']:.2f}"),
        ]
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35

    def _create_detail_sheet(self, ws):
        """创建偏差明细工作表"""
        from openpyxl.styles import Font, PatternFill

        # 获取所有偏差记录
        deviations = self.analytics.base_queryset.select_related().all()

        # 表头
        headers = ['偏差时间', '类型', '系统建议', '实际执行', '验证结果',
                   '盈亏差异', '验证时间', '备注']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # 数据行
        for row, dev in enumerate(deviations, 2):
            ws.cell(row=row, column=1, value=dev.timestamp.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=2, value=dev.get_deviation_type_display())
            ws.cell(row=row, column=3, value=f"{dev.system_action} {dev.system_target}")
            ws.cell(row=row, column=4, value=f"{dev.actual_action} {dev.actual_target}")
            ws.cell(row=row, column=5, value=dev.get_verification_result_display())
            ws.cell(row=row, column=6, value=float(dev.pnl_difference or 0))
            ws.cell(row=row, column=7,
                   value=dev.verified_at.strftime('%Y-%m-%d') if dev.verified_at else '')
            ws.cell(row=row, column=8, value=dev.note or '')

        # 调整列宽
        for col in range(1, 9):
            ws.column_dimensions[chr(64 + col)].width = 18

    def _create_type_stats_sheet(self, ws):
        """创建类型统计工作表"""
        from openpyxl.chart import BarChart, Reference

        # 获取按类型统计的数据
        type_stats = self.analytics.get_intuition_score_by_type()

        # 表头
        headers = ['偏差类型', '总次数', '已验证', '正确次数', '准确率(%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # 数据行
        for row, stat in enumerate(type_stats, 2):
            ws.cell(row=row, column=1, value=stat['deviation_type_display'])
            ws.cell(row=row, column=2, value=stat['total'])
            ws.cell(row=row, column=3, value=stat['verified'])
            ws.cell(row=row, column=4, value=stat['correct'])
            ws.cell(row=row, column=5, value=stat['accuracy_rate'])

        # 创建图表
        chart = BarChart()
        chart.type = 'col'
        chart.title = '各类型偏差准确率'
        chart.y_axis.title = '准确率(%)'
        chart.x_axis.title = '偏差类型'

        data = Reference(ws, min_col=5, min_row=1, max_row=len(type_stats) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(type_stats) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, 'G2')

        # 调整列宽
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_intuition_sheet(self, ws):
        """创建直觉分析工作表"""
        if not self.intuition_tracker:
            ws['A1'] = '未指定用户，无法生成直觉分析'
            return

        # 获取直觉画像
        profile = self.intuition_tracker.get_intuition_profile()

        # 标题
        ws['A1'] = '直觉决策分析报告'
        ws['A1'].font = Font(size=16, bold=True)

        # 基本信息
        ws['A3'] = '总决策数:'
        ws['B3'] = profile['total_decisions']
        ws['A4'] = '正确数:'
        ws['B4'] = profile['correct_count']
        ws['A5'] = '错误数:'
        ws['B5'] = profile['wrong_count']
        ws['A6'] = '准确率:'
        ws['B6'] = f"{profile['accuracy_rate']:.2f}%"

        # 置信区间
        if profile['confidence_level']:
            ws['A8'] = '置信区间 (95%):'
            ws['B8'] = f"{profile['confidence_level']['lower']:.2f}% - {profile['confidence_level']['upper']:.2f}%"

        # 类型细分
        row = 10
        ws[f'A{row}'] = '各类型准确率'
        ws[f'A{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = '偏差类型'
        ws[f'B{row}'] = '总次数'
        ws[f'C{row}'] = '正确次数'
        ws[f'D{row}'] = '准确率(%)'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)

        row += 1
        for type_data in profile['type_breakdown']:
            ws[f'A{row}'] = type_data['deviation_type_display']
            ws[f'B{row}'] = type_data['total']
            ws[f'C{row}'] = type_data['correct']
            ws[f'D{row}'] = type_data['accuracy_rate']
            row += 1

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15


class PDFDeviationReport(DeviationReportGenerator):
    """
    PDF偏差报告生成器
    需要安装 reportlab: pip install reportlab
    """

    def __init__(self, user=None, start_date=None, end_date=None):
        super().__init__(user, start_date, end_date)
        self.has_reportlab = self._check_reportlab()

    def _check_reportlab(self) -> bool:
        """检查是否安装了reportlab"""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            return True
        except ImportError:
            return False

    def generate(self) -> io.BytesIO:
        """
        生成PDF格式的偏差报告

        Returns:
            包含PDF文件的BytesIO对象
        """
        if not self.has_reportlab:
            raise ImportError(
                "生成PDF报告需要安装reportlab。"
                "请运行: pip install reportlab"
            )

        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        # 创建PDF文档
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # 居中
        )

        # 构建内容
        story = []

        # 封面
        story.append(Paragraph("L5偏差日志分析报告", title_style))
        story.append(Spacer(1, 0.3 * inch))

        # 报告信息
        meta = self._get_report_metadata()
        info_data = [
            ['报告期间', f"{meta['date_range']['start']} 至 {meta['date_range']['end']}"],
            ['生成时间', meta['generated_at'].strftime('%Y-%m-%d %H:%M:%S')],
            ['用户', meta['user']],
        ]
        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.5 * inch))

        # 核心指标
        overview = self.analytics.get_overview_stats()
        story.append(Paragraph("核心指标", styles['Heading2']))
        story.append(Spacer(1, 0.2 * inch))

        metrics_data = [
            ['指标', '数值'],
            ['总偏差数', overview['total_deviations']],
            ['已验证数', overview['verification']['verified_count']],
            ['判断正确数', overview['verification']['correct_count']],
            ['直觉准确率', f"{overview['intuition_accuracy']['accuracy_rate']}%"],
            ['盈亏差异总计', f"¥{overview['pnl_stats']['total_pnl_diff']:.2f}"],
        ]
        metrics_table = Table(metrics_data, colWidths=[3 * inch, 3 * inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(metrics_table)

        # 构建PDF
        doc.build(story)
        output.seek(0)
        return output


class DeviationReportService:
    """
    偏差报告服务类
    提供便捷的报告生成接口
    """

    @staticmethod
    def generate_excel_report(user=None, start_date=None, end_date=None) -> io.BytesIO:
        """
        生成Excel偏差报告

        Args:
            user: 指定用户（None表示所有用户）
            start_date: 报告开始日期
            end_date: 报告结束日期

        Returns:
            包含Excel文件的BytesIO对象
        """
        generator = ExcelDeviationReport(user, start_date, end_date)
        return generator.generate()

    @staticmethod
    def generate_pdf_report(user=None, start_date=None, end_date=None) -> io.BytesIO:
        """
        生成PDF偏差报告

        Args:
            user: 指定用户（None表示所有用户）
            start_date: 报告开始日期
            end_date: 报告结束日期

        Returns:
            包含PDF文件的BytesIO对象
        """
        generator = PDFDeviationReport(user, start_date, end_date)
        return generator.generate()

    @staticmethod
    def generate_monthly_report(user, year: int, month: int, format: str = 'excel') -> io.BytesIO:
        """
        生成月度偏差报告

        Args:
            user: 用户对象
            year: 年份
            month: 月份
            format: 报告格式 ('excel' 或 'pdf')

        Returns:
            报告文件的BytesIO对象
        """
        from calendar import monthrange

        # 计算月份的起止日期
        _, last_day = monthrange(year, month)
        start_date = datetime(year, month, 1).date()
        end_date = datetime(year, month, last_day).date()

        if format.lower() == 'pdf':
            return DeviationReportService.generate_pdf_report(user, start_date, end_date)
        else:
            return DeviationReportService.generate_excel_report(user, start_date, end_date)


# 便捷的导出函数
def export_deviations_to_excel(queryset, filename: str = None) -> io.BytesIO:
    """
    将偏差记录导出为Excel

    Args:
        queryset: DeviationLog查询集
        filename: 文件名（可选）

    Returns:
        Excel文件的BytesIO对象
    """
    try:
        import pandas as pd
    except ImportError:
        raise ImportError("请安装 pandas: pip install pandas")

    # 准备数据
    data = []
    for dev in queryset.select_related().all():
        data.append({
            '偏差时间': dev.timestamp,
            '类型': dev.get_deviation_type_display(),
            '系统建议': f"{dev.system_action} {dev.system_target}",
            '实际执行': f"{dev.actual_action} {dev.actual_target}",
            '验证结果': dev.get_verification_result_display(),
            '盈亏差异': float(dev.pnl_difference or 0),
            '验证时间': dev.verified_at,
            '备注': dev.note,
        })

    # 创建DataFrame
    df = pd.DataFrame(data)

    # 写入Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='偏差记录', index=False)

    output.seek(0)
    return output
